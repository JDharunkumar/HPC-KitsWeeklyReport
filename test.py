import os
from datetime import datetime
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import smtplib
from email.message import EmailMessage

# ================================
# 1. SQL Server Configuration
# ================================
DB_SERVER   = 'STLV-SQLPROD04' #'BGLV-SQLDEV01'
DB_DATABASE = 'RuleBuilderHSMN'

# ================================
# 2. Email Configuration
# ================================
SMTP_SERVER = 'smtp-hsm.internal.hussmann.com'
SMTP_PORT   = 25
EMAIL_TO    = [
        'Monica.Rosales@hussmann.com',
        'Ingrid.Arellanes@hussmann.com'
]
EMAIL_CC    = [
    'gregory.t.garnier@hussmann.com',
    'Dhiraj.Paswan@hussmann.com',
    'Dharunkumar.Jayabal@hussmann.com' 

]

# ================================
# 3. Inline SQL Queries
# ================================
MODELS_QUERY = """
SELECT 
    pl.Name            AS ProductLineName,
    tpm.Name,
    tpm.Description,
    tpm.CurrentListPrice,
    tpm.stdCost,
    FORMAT(tpm.DateCreated, 'MM/dd/yyyy')  AS DateCreated,
    FORMAT(tpm.DateModified, 'MM/dd/yyyy') AS DateModified,
    tpm.SiteName
FROM tProductModel tpm WITH (NOLOCK)
INNER JOIN tProductLine pl WITH (NOLOCK)
    ON tpm.DBID_ProductLine = pl.DBID
WHERE tpm.Remove <> 1
ORDER BY pl.Name
"""

OPTIONS_QUERY = """
SELECT 
    pl.Name            AS ProductLineName,
    p.Name             AS ProductName,
    o.Name             AS OptionName,
    o.Description,
    o.CurrentListPrice,
    o.stdCost,
    FORMAT(o.DateCreated, 'MM/dd/yyyy')  AS DateCreated,
    FORMAT(o.DateModified, 'MM/dd/yyyy') AS DateModified,
    o.Notes,
    f.CategoryTag
FROM tPLPFO pf WITH (NOLOCK)
INNER JOIN tOption o WITH (NOLOCK)
    ON pf.DBID_ProductLine = o.DBID_ProductLine
   AND pf.DBID_Option      = o.DBID
   AND pf.DBID_PLRev       = o.DBID_PLRev
INNER JOIN tFeature f WITH (NOLOCK)
    ON pf.DBID_Feature     = f.DBID
   AND pf.DBID_ProductLine = f.DBID_ProductLine
   AND pf.DBID_PLRev       = f.DBID_PLRev
INNER JOIN tProductLine pl WITH (NOLOCK)
    ON pf.DBID_ProductLine = pl.DBID
INNER JOIN tProduct p WITH (NOLOCK)
    ON pf.DBID_ProductLine = p.DBID_ProductLine
   AND pf.DBID_Product     = p.DBID
WHERE 
    pf.[Remove] <> 1
    AND (o.KitExpiryDate >= CONVERT(char, GETDATE(),101) OR o.KitExpiryDate IS NULL)
ORDER BY pl.Name, o.Name
"""

SPECIALS_QUERY = """
SELECT 
    pl.Name            AS ProductLineName,
    p.Name             AS ProductName,
    o.Name             AS OptionName,
    o.Description,
    o.CurrentListPrice,
    o.stdCost,
    FORMAT(o.DateCreated, 'MM/dd/yyyy')  AS DateCreated,
    FORMAT(o.DateModified, 'MM/dd/yyyy') AS DateModified,
    o.Notes,
    s.CategoryTag
FROM tPLPSO s WITH (NOLOCK)
INNER JOIN tOption o WITH (NOLOCK)
    ON s.DBID_ProductLine = o.DBID_ProductLine
   AND s.DBID_Option      = o.DBID
   AND s.DBID_PLRev       = o.DBID_PLRev
INNER JOIN tProductLine pl WITH (NOLOCK)
    ON s.DBID_ProductLine = pl.DBID
INNER JOIN tProduct p WITH (NOLOCK)
    ON s.DBID_ProductLine = p.DBID_ProductLine
   AND s.DBID_Product     = p.DBID
WHERE 
    s.[Remove] <> 1
    AND (o.KitExpiryDate >= CONVERT(char, GETDATE(),101) OR o.KitExpiryDate IS NULL)
ORDER BY pl.Name, o.Name
"""

# ================================
# 4. Database Helper
# ================================
def get_db_engine():
    conn_str = (
        f"mssql+pyodbc://{DB_SERVER}/{DB_DATABASE}"
        "?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
    )
    return create_engine(conn_str)

def fetch_data(product_line_id=None, product_id=None):
    engine = get_db_engine()
    
    # Add filters if provided
    def add_filters(sql, alias_pl='pl.DBID', alias_p='p.DBID'):
        if product_line_id:
            sql += f" AND {alias_pl} = {product_line_id}"
        if product_id:
            sql += f" AND {alias_p} = {product_id}"
        return sql

    # Fetch Models data
    models_query = MODELS_QUERY
    if product_line_id:
        models_query = models_query.replace("ORDER BY pl.Name", f"AND pl.DBID = {product_line_id} ORDER BY pl.Name")
    models_df = pd.read_sql(models_query, engine)
    
    # Fetch Options and Specials data
    options_df = pd.read_sql(add_filters(OPTIONS_QUERY), engine)
    specials_df = pd.read_sql(add_filters(SPECIALS_QUERY), engine)
    
    return models_df, options_df, specials_df

# ================================
# 5. Excel Generation with Two Sheets
# ================================
def create_excel_report(product_line_id=None, product_id=None):
    models_df, options_df, specials_df = fetch_data(product_line_id, product_id)
    
    # Prepare Kits data (Options + Specials)
    options_df['Type'] = 'Option'
    specials_df['Type'] = 'Special'
    
    # Add missing Notes column to specials if not present
    if 'Notes' not in specials_df.columns:
        specials_df['Notes'] = ''
    
    # Combine options and specials
    kits_df = pd.concat([options_df, specials_df], ignore_index=True)
    
    # Format currency for both sheets
    def format_currency(x):
        try:
            return f"${float(x):,.2f}"
        except:
            return "" if pd.isna(x) else str(x)
    
    # Format Kits data
    kits_rows = []
    for _, row in kits_df.iterrows():
        kits_rows.append({
            'ProductLine Name': row['ProductLineName'],
            'Option Name': row['OptionName'],
            'Description': row['Description'],
            'Date Created': row['DateCreated'],
            'Date Modified': row['DateModified'],
            'Current List Price': format_currency(row['CurrentListPrice']),
            'Standard Cost': format_currency(row['stdCost'])
        })
    
    # Format Models data
    models_rows = []
    for _, row in models_df.iterrows():
        models_rows.append({
            'ProductLine Name': row['ProductLineName'],
            'Model Name': row['Name'],
            'Description': row['Description'],
            'Date Created': row['DateCreated'],
            'Date Modified': row['DateModified'],
            'Current List Price': format_currency(row['CurrentListPrice']),
            'Standard Cost': format_currency(row['stdCost']),
            'Site Name': row['SiteName']
        })
    
    # Create DataFrames
    kits_final_df = pd.DataFrame(kits_rows)
    models_final_df = pd.DataFrame(models_rows)
    
    # Sort data
    kits_final_df = kits_final_df.sort_values(['ProductLine Name', 'Option Name'], ascending=[True, True])
    models_final_df = models_final_df.sort_values(['ProductLine Name', 'Model Name'], ascending=[True, True])
    
    # Reset index
    kits_final_df = kits_final_df.reset_index(drop=True)
    models_final_df = models_final_df.reset_index(drop=True)
    
    # Create Excel file
    os.makedirs("reports", exist_ok=True)
    file_name = f"reports/UniverseKits And Model Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        kits_final_df.to_excel(writer, sheet_name='Kits', index=False)
        models_final_df.to_excel(writer, sheet_name='Models', index=False)
    
    # Format Excel sheets
    wb = load_workbook(file_name)
    
    # Format both sheets
    for sheet_name in ['Kits', 'Models']:
        ws = wb[sheet_name]
        
        # Define font styles
        header_font = Font(bold=True, size=11)
        data_font = Font(size=8)
        center_alignment = Alignment(vertical="center")
        
        # Header style (row 1)
        header_fill = PatternFill("solid", fgColor="FFD700")
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = header_fill
        
        # Data style (rows 2 onwards)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font
                cell.alignment = center_alignment
        
        # Auto-fit columns
        for col in ws.columns:
            w = max(len(str(c.value)) for c in col if c.value) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, 50)
        
        # Add auto filter
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze the first row
        ws.freeze_panes = 'A2'
    
    wb.save(file_name)
    return file_name

# ================================
# 6. Email Sender
# ================================
def send_email(file_path):
    msg = EmailMessage()
    msg['Subject'] = f'Kit Usage Report ({datetime.now():%d-%b-%Y})'
    msg['From'] = 'reports@hussmann.com'
    msg['To'] = ','.join(EMAIL_TO)
    
    # Add CC recipients if they exist
    if 'EMAIL_CC' in globals() and EMAIL_CC:
        msg['Cc'] = ', '.join(EMAIL_CC)
    
    # Create HTML content with blue disclaimer
    html_content = """
    <html>
    <body>
        <p>Hello,</p>
        <p>Please find attached the UniverseKits And Model Report</p>
        <p>Regards,<br>
        Automated Reporting System</p>
        <p>This is an automated email, please do not reply.</p>
        <p><span style="color: blue;"><strong>Disclaimer:</strong><br>
        The content of this email is intended solely for the person or entity to whom it is addressed. It may contain confidential information. If you are not the intended recipient or have received this message in error, please contact HPC Support immediately.</span></p>
    </body>
    </html>
    """
    
    # Set HTML content
    msg.set_content(html_content, subtype='html')
    
    with open(file_path, 'rb') as f:
        msg.add_attachment(
            f.read(),
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=os.path.basename(file_path)
        )
    
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
        s.send_message(msg)

# ================================
# 7. Main
# ================================
if __name__ == '__main__':
    try:
        path = create_excel_report()
        #send_email(path)
        print("Report generated:", path)
    except Exception as e:
        print("Error:", e)
